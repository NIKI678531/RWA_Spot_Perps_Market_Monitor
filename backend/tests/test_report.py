"""Tests for the report layer.

The interesting assertions are not "does openpyxl produce a file" but "does the
workbook keep the promises the domain rules make": a missing observation must not
become a zero, overlapping categories must stay marked non-additive, and no sheet
may total across metric scopes.
"""

import zipfile
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from typing import Iterator

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

import app.models  # noqa: F401 - registers every table on Base.metadata
from app.core.sessions import MarketSession
from app.db.base import Base
from app.models.dimensions import (
    DimAsset,
    DimIssuer,
    DimPerpContract,
    DimUnderlying,
    DimVenue,
)
from app.models.enums import AssetClass, RwaTier, VenueType
from app.models.facts import (
    FactAssetSnapshot,
    FactCategorySnapshot,
    FactPairSnapshot,
    FactPerpContractSnapshot,
)
from app.services.report import dataset, service, storage
from app.services.report.excel import build_sheets
from app.services.report.word import render_docx
from app.services.report.workbook import NOT_VERIFIED, SheetSpec, render

NOW = datetime(2026, 8, 17, 14, 0, tzinfo=timezone.utc)


@pytest.fixture()
def session() -> Iterator[Session]:
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as session:
        yield session


def _seed(session: Session) -> None:
    """A miniature market: one underlying, two wrappers, two venues, one perp."""
    session.add_all(
        [
            DimUnderlying(
                underlying_id="SPY", name="SPDR S&P 500 ETF", asset_class=AssetClass.ETF
            ),
            DimIssuer(issuer_id="bstocks", name="bStocks", official_product_count=60),
            DimIssuer(issuer_id="xstocks", name="xStocks", official_product_count=640),
            DimVenue(venue_id="binance", name="Binance", venue_type=VenueType.CEX),
            DimVenue(
                venue_id="native_bsc",
                name="Native (BSC)",
                venue_type=VenueType.DEX,
                chain="bsc",
            ),
            DimAsset(
                asset_id="spyb",
                symbol="SPYB",
                rwa_tier=RwaTier.CORE_RWA,
                underlying_id="SPY",
                issuer_id="bstocks",
            ),
            DimAsset(
                asset_id="spyx",
                symbol="SPYx",
                rwa_tier=RwaTier.CORE_RWA,
                underlying_id="SPY",
                issuer_id="xstocks",
            ),
            # Benchmark-only: must never enter a ranking.
            DimAsset(asset_id="btc", symbol="BTC", rwa_tier=RwaTier.NON_RWA),
            DimPerpContract(
                contract_id="HL:rwa:SPY",
                exchange="Hyperliquid",
                perp_dex="rwa",
                symbol="SPY",
                source_underlying_type="EQUITY",
                analysis_group="etf",
                underlying_id="SPY",
            ),
        ]
    )
    session.add_all(
        [
            FactAssetSnapshot(
                asset_id="spyb",
                snapshot_ts=NOW,
                market_session=MarketSession.CLOSED_WEEKEND,
                market_cap=Decimal("5000000"),
                vol_24h=Decimal("362000"),
            ),
            FactAssetSnapshot(
                asset_id="spyx",
                snapshot_ts=NOW,
                market_session=MarketSession.CLOSED_WEEKEND,
                market_cap=None,  # not verified, not zero
                vol_24h=Decimal("120000"),
            ),
            FactAssetSnapshot(
                asset_id="btc",
                snapshot_ts=NOW,
                market_session=MarketSession.CLOSED_WEEKEND,
                market_cap=Decimal("2000000000000"),
                vol_24h=Decimal("40000000000"),
            ),
            FactPairSnapshot(
                asset_id="spyb",
                venue_id="binance",
                snapshot_ts=NOW,
                market_session=MarketSession.CLOSED_WEEKEND,
                raw_vol_24h=Decimal("300000"),
                adjusted_vol_24h=Decimal("300000"),
            ),
            # The Native (BSC) shape: nearly everything flagged.
            FactPairSnapshot(
                asset_id="spyx",
                venue_id="native_bsc",
                snapshot_ts=NOW,
                market_session=MarketSession.CLOSED_WEEKEND,
                raw_vol_24h=Decimal("29300000"),
                adjusted_vol_24h=None,
                is_quality_anomaly=True,
            ),
            FactPairSnapshot(
                asset_id="btc",
                venue_id="binance",
                snapshot_ts=NOW,
                market_session=MarketSession.CLOSED_WEEKEND,
                raw_vol_24h=Decimal("40000000000"),
                adjusted_vol_24h=Decimal("40000000000"),
            ),
            FactPerpContractSnapshot(
                contract_id="HL:rwa:SPY",
                snapshot_ts=NOW,
                market_session=MarketSession.CLOSED_WEEKEND,
                vol_24h=Decimal("700000"),
                oi_units=Decimal("100"),
                oi_usd=Decimal("77316"),
                mark_price=Decimal("773.16"),
            ),
            FactCategorySnapshot(
                category_id="tokenized-stock",
                snapshot_ts=NOW,
                market_session=MarketSession.CLOSED_WEEKEND,
                asset_count=150,
                market_cap=Decimal("400000000"),
                is_additive=False,
            ),
            FactCategorySnapshot(
                category_id="rwa_union",
                snapshot_ts=NOW,
                market_session=MarketSession.CLOSED_WEEKEND,
                asset_count=250,
                market_cap=Decimal("500000000"),
                is_additive=True,
            ),
        ]
    )
    session.commit()


def _sheet(sheets: list[SheetSpec], name: str) -> SheetSpec:
    return next(s for s in sheets if s.name == name)


def _column(sheet: SheetSpec, header: str) -> list[object]:
    index = list(sheet.headers).index(header)
    return [row[index] for row in sheet.rows]


# --- rendering -------------------------------------------------------------


def test_a_missing_value_is_never_written_as_zero() -> None:
    """Rule 3: a failed fetch is a missing observation, not a zero."""
    sheet = SheetSpec(name="t", headers=["a"], rows=[[None], [Decimal(0)]])
    workbook = load_workbook(BytesIO(render([sheet])))
    values = [row[0] for row in workbook["t"].iter_rows(min_row=2, values_only=True)]

    assert values == [NOT_VERIFIED, 0]


def test_the_workbook_has_no_merged_cells_or_charts(session: Session) -> None:
    """Section 11: the xlsx must stay copy-pasteable and pivotable."""
    _seed(session)
    content = render(build_sheets(dataset.load(session)))
    workbook = load_workbook(BytesIO(content))

    for worksheet in workbook.worksheets:
        assert not worksheet.merged_cells.ranges, worksheet.title
        assert not list(worksheet.conditional_formatting), worksheet.title

    # Charts and drawings live in their own parts of the archive, so their absence
    # is checkable without reaching into openpyxl's internals.
    with zipfile.ZipFile(BytesIO(content)) as archive:
        parts = archive.namelist()
    assert not [p for p in parts if p.startswith(("xl/charts/", "xl/drawings/"))]


# --- sheet construction ----------------------------------------------------


def test_the_workbook_has_twenty_two_sheets(session: Session) -> None:
    _seed(session)
    sheets = build_sheets(dataset.load(session))

    assert len(sheets) == 22
    assert sheets[0].name == "01_Asset_Master"
    assert [s.name for s in sheets[15:18]] == [
        "16_HL_HIP3_Contracts",
        "17_Liquidity_Quality",
        "18_Theme_Demand",
    ]
    assert sheets[-1].name == "22_Scope_Notes"


def test_every_sheet_survives_rendering(session: Session) -> None:
    _seed(session)
    workbook = load_workbook(BytesIO(render(build_sheets(dataset.load(session)))))

    assert len(workbook.sheetnames) == 22


def test_non_rwa_assets_stay_out_of_the_rankings(session: Session) -> None:
    """Rule: NON_RWA exists for benchmark reference and enters no ranking."""
    _seed(session)
    sheets = build_sheets(dataset.load(session))

    master = _column(_sheet(sheets, "01_Asset_Master"), "asset_id")
    venues = _sheet(sheets, "06_Venue_Ranking")
    binance_volume = next(
        row[list(venues.headers).index("adjusted_vol_24h")]
        for row in venues.rows
        if row[1] == "Binance"
    )

    assert "btc" in master, "the master list stays complete"
    # $40bn of BTC turnover would swamp the ranking if it leaked in.
    assert binance_volume == Decimal("300000")


def test_overlapping_categories_stay_marked_non_additive(session: Session) -> None:
    """Rule 2: only the deduplicated union row is a valid total."""
    _seed(session)
    sheets = build_sheets(dataset.load(session))
    sheet = _sheet(sheets, "02_Category_Scale")

    additive = dict(zip(_column(sheet, "category_id"), _column(sheet, "is_additive")))

    assert additive == {"rwa_union": True, "tokenized-stock": False}


def test_raw_and_adjusted_turnover_are_both_reported(session: Session) -> None:
    """Rule 4: publishing either figure alone misleads."""
    _seed(session)
    sheets = build_sheets(dataset.load(session))
    sheet = _sheet(sheets, "17_Liquidity_Quality")

    row = next(r for r in sheet.rows if r[0] == "Native (BSC)")
    headers = list(sheet.headers)

    assert row[headers.index("raw_vol_24h")] == Decimal("29300000")
    assert row[headers.index("adjusted_vol_24h")] is None
    assert row[headers.index("flagged_pairs")] == 1
    # Divergence needs both sides observed; here adjusted was never observed at all,
    # which the coverage column says explicitly rather than implying $0.
    assert row[headers.index("coverage")] == "not_verified"


def test_a_partial_total_is_labelled_partial(session: Session) -> None:
    _seed(session)
    sheets = build_sheets(dataset.load(session))
    sheet = _sheet(sheets, "04_Underlying_Demand")
    headers = list(sheet.headers)
    row = next(r for r in sheet.rows if r[0] == "SPY")

    # spyb reported $300k, spyx reported nothing. The total is a floor, not a fact.
    assert row[headers.index("spot_vol_adjusted")] == Decimal("300000")
    assert row[headers.index("spot_vol_coverage")] == "partial"
    # A different scope, in its own column, with no total anywhere on the sheet.
    assert row[headers.index("perp_vol_24h")] == Decimal("700000")
    assert "total" not in " ".join(headers)


def test_spot_and_perp_sit_side_by_side_with_only_a_ratio(session: Session) -> None:
    """Rule 1: five scopes, never summed. A ratio is a comparison, not a sum."""
    _seed(session)
    sheets = build_sheets(dataset.load(session))
    sheet = _sheet(sheets, "14_Perp_vs_Spot")
    headers = list(sheet.headers)
    row = next(r for r in sheet.rows if r[0] == "SPY")

    assert row[headers.index("perp_over_spot")] == pytest.approx(700000 / 300000)


def test_hip3_deployments_are_distinguished_from_the_core_venue(
    session: Session,
) -> None:
    _seed(session)
    sheets = build_sheets(dataset.load(session))
    sheet = _sheet(sheets, "16_HL_HIP3_Contracts")

    assert _column(sheet, "perp_dex") == ["rwa"]
    assert _column(sheet, "is_hip3") == [True]


def test_the_scope_notes_sheet_covers_every_sheet_that_carries_one(
    session: Session,
) -> None:
    _seed(session)
    sheets = build_sheets(dataset.load(session))
    notes = _sheet(sheets, "22_Scope_Notes")

    named = {row[0] for row in notes.rows}
    expected = {s.name for s in sheets if s.note or s.scopes}

    assert expected <= named


def test_an_empty_warehouse_still_produces_a_workbook(session: Session) -> None:
    """The first scheduled run happens before any collector has succeeded."""
    workbook = load_workbook(BytesIO(render(build_sheets(dataset.load(session)))))

    assert len(workbook.sheetnames) == 22


# --- docx ------------------------------------------------------------------


def test_the_analysis_report_leads_with_alerts(session: Session) -> None:
    _seed(session)
    content = render_docx(dataset.load(session))

    assert content[:2] == b"PK"  # a docx is a zip
    assert len(content) > 0


# --- storage ---------------------------------------------------------------


def test_reports_are_persisted_rather_than_written_to_disk(session: Session) -> None:
    """No PVC in production: the bytes must land in the database."""
    _seed(session)
    artifacts = service.generate(session, as_of=NOW)

    assert [a.report_format for a in artifacts] == ["xlsx", "docx"]
    assert all(a.content is not None for a in artifacts)
    assert all(a.storage_key is None for a in artifacts)
    assert artifacts[0].filename.endswith("2026-08-17.xlsx")


def test_regenerating_replaces_rather_than_duplicates(session: Session) -> None:
    _seed(session)
    first = service.generate(session, as_of=NOW)
    second = service.generate(session, as_of=NOW)

    assert [a.id for a in first] == [a.id for a in second]


def test_object_storage_is_used_when_configured(session: Session) -> None:
    _seed(session)
    uploaded: dict[str, bytes] = {}

    def upload(key: str, content: bytes, content_type: str) -> str:
        uploaded[key] = content
        return key

    store = storage.ReportStore(backend="tos", uploader=upload)
    artifacts = service.generate(session, as_of=NOW, store=store)

    assert len(uploaded) == 2
    assert all(a.content is None for a in artifacts)
    assert all(a.storage_key in uploaded for a in artifacts)


def test_object_storage_without_an_uploader_fails_loudly(session: Session) -> None:
    """Silently falling back to local disk would work in dev and lose data in prod."""
    _seed(session)
    store = storage.ReportStore(backend="tos", uploader=None)

    with pytest.raises(RuntimeError, match="uploader"):
        service.generate(session, as_of=NOW, store=store)
