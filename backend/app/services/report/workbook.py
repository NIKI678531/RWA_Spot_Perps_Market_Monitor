"""Plain workbook rendering.

ARCHITECTURE.md §11 requires the xlsx stay plain: no conditional formatting, no
embedded charts, no merged cells. That is not minimalism for its own sake — the
workbook exists to be copied out of and pivoted elsewhere, and each of those three
features breaks a paste. Visualisation lives on the web side.

The one presentational decision made here is how a missing observation renders. It
is written as the text ``Not verified``, never as a blank and never as 0. A blank
cell reads as zero once it lands in a SUM, and a zero asserts that nobody traded
when what actually happened is that we failed to look.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from enum import Enum
from io import BytesIO
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

#: What a null money column renders as. See the module docstring.
NOT_VERIFIED = "Not verified"

#: Excel's own limit on a sheet name.
_MAX_SHEET_NAME = 31

_HEADER_FONT = Font(bold=True)
_MIN_WIDTH = 10
_MAX_WIDTH = 42


@dataclass(frozen=True, slots=True)
class SheetSpec:
    """One rectangular sheet.

    ``note`` does not go into the grid. Prepending a note row would shift the header
    off row 1 and break every downstream pivot; the notes are collected into their
    own sheet instead.
    """

    name: str
    headers: Sequence[str]
    rows: Sequence[Sequence[Any]]
    note: str = ""
    #: Metric scopes this sheet reports, for the scope-notes sheet. A sheet listing
    #: more than one is stating that its columns are side by side and not addable.
    scopes: Sequence[str] = field(default_factory=tuple)

    def __post_init__(self) -> None:
        if len(self.name) > _MAX_SHEET_NAME:
            raise ValueError(f"sheet name too long for Excel: {self.name!r}")


def to_cell(value: Any) -> Any:
    """Coerce one Python value to something openpyxl can write.

    ``None`` becomes the ``Not verified`` marker rather than an empty cell, which is
    the whole reason this function exists rather than passing values through.
    """
    if value is None:
        return NOT_VERIFIED
    if isinstance(value, bool):
        # Checked before Decimal/int: bool is a subclass of int and would otherwise
        # render as 1/0, which reads as a quantity.
        return "Yes" if value else "No"
    if isinstance(value, Decimal):
        # Excel stores float64 regardless, so the precision is lost at write time
        # either way. Converting here keeps the cell numeric instead of text.
        return float(value)
    if isinstance(value, Enum):
        return str(value.value)
    if isinstance(value, datetime):
        # Excel has no timezone concept; a tz-aware datetime raises on write.
        return value.replace(tzinfo=None)
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def render(sheets: Sequence[SheetSpec]) -> bytes:
    """Render sheets to xlsx bytes.

    Returns bytes rather than writing a file: production K8s provides no PVC, so
    anything written to the container filesystem is lost on the next rollout.
    """
    if not sheets:
        raise ValueError("a workbook needs at least one sheet")

    workbook = Workbook()
    # Workbook() ships with one sheet already; remove it so sheet order is exactly
    # the order given.
    workbook.remove(workbook.worksheets[0])

    for spec in sheets:
        worksheet = workbook.create_sheet(title=spec.name)
        worksheet.append(list(spec.headers))
        for cell in worksheet[1]:
            cell.font = _HEADER_FONT
        for row in spec.rows:
            worksheet.append([to_cell(v) for v in row])

        # Freezing the header is neither a merge, a chart, nor conditional
        # formatting, and it survives a copy-paste intact.
        worksheet.freeze_panes = "A2"
        _fit_columns(worksheet, spec)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _fit_columns(worksheet: Any, spec: SheetSpec) -> None:
    """Set column widths from the header and the first rows.

    Only the first rows are sampled: scanning a 5,000-row sheet to widen a column by
    two characters costs more than it is worth.
    """
    sample = list(spec.rows[:50])
    for index, header in enumerate(spec.headers, start=1):
        widest = len(str(header))
        for row in sample:
            if index - 1 < len(row):
                widest = max(widest, len(str(to_cell(row[index - 1]))))
        width = min(max(widest + 2, _MIN_WIDTH), _MAX_WIDTH)
        worksheet.column_dimensions[get_column_letter(index)].width = width
